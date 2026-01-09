from flask import Flask, render_template, request, jsonify, send_file
from werkzeug.utils import secure_filename
import requests
from bs4 import BeautifulSoup
import re
import time
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import json
import io
from openpyxl import Workbook, load_workbook
# Removed styling imports - keeping it plain
# Removed get_column_letter import - not needed for plain export
from datetime import datetime
import os
import tempfile
import hashlib
from functools import lru_cache
from threading import Lock
import pickle
import concurrent.futures
from concurrent.futures import ThreadPoolExecutor, as_completed
import psutil  # For CPU monitoring
import time
import random
from threading import Event
import diskcache as dc

# Playwright removed - using Selenium and requests only

app = Flask(__name__)

# PERFORMANCE: Caching system with diskcache
CACHE_DIR = "cache"
if not os.path.exists(CACHE_DIR):
    os.makedirs(CACHE_DIR)

# Replace the entire ProfileCache class with diskcache
profile_cache = dc.Cache(CACHE_DIR,
                         size_limit=int(4e9),  # 4 GB disk cache limit
                         eviction_policy='least-recently-used',
                         cull_limit=2000,  # Number of items to cull when limit is reached
                         )

# PERFORMANCE: WebDriver Pool for concurrent processing
from queue import Queue
import threading

class WebDriverPool:
    def __init__(self, pool_size=10):
        self.pool = Queue(maxsize=pool_size)
        self.pool_size = pool_size
        self.drivers_created = 0
        self.pool_lock = Lock()
        
        # LAZY LOADING: Don't pre-create drivers - create them on demand
        # This enables instant startup while maintaining functionality
        print(f"WebDriver pool initialized (lazy loading, max {pool_size} drivers)")
    
    def _is_driver_valid(self, driver):
        """Check if driver session is still valid"""
        try:
            if not driver:
                return False
            # Try a simple operation to test session validity
            driver.current_url
            return True
        except Exception:
            return False
    
    def _create_driver(self):
        try:
            chrome_options = Options()
            chrome_options.add_argument('--headless')
            chrome_options.add_argument('--no-sandbox')
            chrome_options.add_argument('--disable-dev-shm-usage')
            chrome_options.add_argument('--disable-gpu')
            chrome_options.add_argument('--disable-blink-features=AutomationControlled')
            chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
            chrome_options.add_experimental_option('useAutomationExtension', False)
            chrome_options.add_argument('--window-size=1920,1080')
            chrome_options.add_argument('--disable-web-security')
            chrome_options.add_argument('--allow-running-insecure-content')
            chrome_options.add_argument('--disable-logging')
            chrome_options.add_argument('--log-level=3')
            chrome_options.add_argument('--disable-extensions')
            chrome_options.add_argument('--disable-plugins')
            chrome_options.add_argument('--disable-images')
            # Removed --disable-javascript as it breaks TikTok functionality
            chrome_options.add_argument('--disable-css')
            chrome_options.add_argument('--enable-unsafe-swiftshader')
            chrome_options.add_argument('--disable-webgl')
            chrome_options.add_argument('--disable-accelerated-2d-canvas')
            chrome_options.add_argument('--no-first-run')
            chrome_options.add_argument('--disable-default-apps')
            chrome_options.add_argument('--disable-background-timer-throttling')
            chrome_options.add_argument('--disable-backgrounding-occluded-windows')
            chrome_options.add_argument('--disable-renderer-backgrounding')
            
            # Additional speed optimizations
            chrome_options.add_argument('--disable-background-networking')
            chrome_options.add_argument('--disable-sync')
            chrome_options.add_argument('--disable-translate')
            chrome_options.add_argument('--disable-features=TranslateUI')
            chrome_options.add_argument('--disable-features=VizDisplayCompositor')
            chrome_options.add_argument('--disable-hang-monitor')
            chrome_options.add_argument('--disable-prompt-on-repost')
            chrome_options.add_argument('--disable-domain-reliability')
            chrome_options.add_argument('--aggressive-cache-discard')
            chrome_options.add_argument('--disable-component-extensions-with-background-pages')
            chrome_options.add_argument('--disable-software-rasterizer')
            chrome_options.add_argument('--disable-features=TranslateUI')
            chrome_options.add_argument('--disable-ipc-flooding-protection')
            chrome_options.add_argument('--hide-scrollbars')
            chrome_options.add_argument('--mute-audio')
            
            # Try multiple methods to create driver
            driver = None
            
            # Method 1: Try with ChromeDriverManager
            try:
                from selenium.webdriver.chrome.service import Service
                service = Service(ChromeDriverManager().install())
                driver = webdriver.Chrome(service=service, options=chrome_options)
                print("WebDriver created with ChromeDriverManager")
            except Exception as e:
                print(f"ChromeDriverManager failed: {e}")
                
                # Method 2: Try without service (system PATH)
                try:
                    driver = webdriver.Chrome(options=chrome_options)
                    print("WebDriver created with system Chrome")
                except Exception as e2:
                    print(f"System Chrome failed: {e2}")
                    
                    # Method 3: Try with different Chrome binary path
                    try:
                        chrome_options.binary_location = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
                        driver = webdriver.Chrome(options=chrome_options)
                        print("WebDriver created with specific Chrome path")
                    except Exception as e3:
                        print(f"Specific Chrome path failed: {e3}")
                        return None
            
            if driver:
                driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
                self.drivers_created += 1
                return driver
            else:
                return None
                
        except Exception as e:
            print(f"Failed to create WebDriver: {e}")
            return None
    
    def get_driver(self, timeout=30):
        try:
            # Try to get existing driver from pool
            driver = self.pool.get(timeout=timeout)
            
            # Validate the driver before returning it
            if self._is_driver_valid(driver):
                return driver
            else:
                # Driver is invalid, close it and create a new one
                print("Driver session invalid, creating new one")
                try:
                    driver.quit()
                except:
                    pass
                return self._create_driver()
        except:
            # LAZY LOADING: Create driver on demand if pool is empty
            if self.drivers_created < self.pool_size:
                driver = self._create_driver()
                if driver:
                    return driver
            
            # If we've reached max drivers, create temporary one
            return self._create_driver()
    
    def return_driver(self, driver):
        if driver:
            try:
                # Try to return driver to pool if there's space
                self.pool.put_nowait(driver)
            except:
                # Pool is full, close the driver
                try:
                    driver.quit()
                except:
                    pass
    
    def close_all(self):
        """Enhanced cleanup method to prevent threading issues"""
        while not self.pool.empty():
            try:
                driver = self.pool.get_nowait()
                driver.quit()
            except:
                break
        
        # Additional cleanup for any lingering Chrome processes
        try:
            import psutil
            for proc in psutil.process_iter(['pid', 'name', 'cmdline']):
                try:
                    if proc.info['name'] and 'chrome' in proc.info['name'].lower():
                        cmdline = proc.info.get('cmdline', [])
                        if cmdline and any('headless' in str(arg).lower() for arg in cmdline):
                            proc.terminate()
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    pass
        except ImportError:
            pass  # psutil not available
        except Exception:
            pass  # Ignore any cleanup errors

# Playwright browser pool removed

# Initialize WebDriver pool with error handling
try:
    webdriver_pool = WebDriverPool(pool_size=20)
    print("WebDriver pool initialized successfully")
except Exception as e:
    print(f"WebDriver pool initialization failed: {e}")
    print("App will run in requests-only mode (no Selenium fallback)")
    webdriver_pool = None

app = Flask(__name__)

# PERFORMANCE: Flask optimization for better threading
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0  # Disable caching for development

# PERFORMANCE: CPU Monitoring System (no throttling)
class CPUMonitor:
    def __init__(self, check_interval=1.0):
        self.check_interval = check_interval
        self.cpu_history = []
        self.max_history = 10  # Keep last 10 CPU readings
        
    def get_cpu_usage(self):
        """Get current CPU usage percentage"""
        try:
            return psutil.cpu_percent(interval=0.1)
        except:
            return 0
    
    def get_cpu_history(self):
        """Get CPU usage history for monitoring"""
        current_cpu = self.get_cpu_usage()
        
        # Add to history
        self.cpu_history.append(current_cpu)
        if len(self.cpu_history) > self.max_history:
            self.cpu_history.pop(0)
        
        return self.cpu_history
    
    def get_optimal_workers(self, base_workers=None):
        """Calculate optimal number of workers based on CPU cores"""
        import os
        cpu_count = os.cpu_count() or 4  # Default to 4 if detection fails
        
        # Use 5 workers maximum for better resource management
        optimal_workers = 20
        
        return optimal_workers

# Initialize CPU monitor
cpu_monitor = CPUMonitor()

class TikTokScraper:
    def __init__(self):
        # User agent rotation for better rate limiting avoidance
        self.user_agents = [
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:121.0) Gecko/20100101 Firefox/121.0',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:121.0) Gecko/20100101 Firefox/121.0',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Edge/120.0.0.0 Safari/537.36',
        ]
        
        self.headers = {
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
            'Accept-Language': 'en-US,en;q=0.9',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
            'Sec-Fetch-Dest': 'document',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-Site': 'none',
            'Cache-Control': 'max-age=0',
            'DNT': '1',
        }
        
        # Request tracking for rate limiting
        self.request_count = 0
        self.last_request_time = 0
        # PERFORMANCE: Shared session for connection reuse
        self.session = requests.Session()
        self.session.headers.update(self.headers)
        # Configure session for optimal performance
        adapter = requests.adapters.HTTPAdapter(
            pool_connections=20,  # Number of connection pools
            pool_maxsize=20,      # Max connections per pool
            max_retries=3,        # Retry failed requests
            pool_block=False      # Don't block when pool is full
        )
        self.session.mount('http://', adapter)
        self.session.mount('https://', adapter)
    
    def get_random_user_agent(self):
        """Get a random user agent for request rotation"""
        return random.choice(self.user_agents)
    
    def add_request_delay(self):
        """Add intelligent delay between requests to avoid rate limiting"""
        current_time = time.time()
        time_since_last = current_time - self.last_request_time
        
        # Base delay between 0.5-2 seconds
        base_delay = random.uniform(0.5, 2.0)
        
        # Increase delay if we're making requests too fast
        if time_since_last < 1.0:
            base_delay += random.uniform(1.0, 3.0)
        
        # Add worker-specific delay to stagger requests
        worker_delay = random.uniform(0.1, 0.5)
        
        total_delay = base_delay + worker_delay
        
        if total_delay > 0:
            print(f"⏳ Rate limiting delay: {total_delay:.2f}s")
            time.sleep(total_delay)
        
        self.last_request_time = time.time()
        self.request_count += 1
    
    def make_request_with_backoff(self, url, max_retries=3):
        """Make request with exponential backoff for rate limit errors"""
        for attempt in range(max_retries):
            try:
                # Add delay before request
                self.add_request_delay()
                
                # Rotate user agent
                headers = self.headers.copy()
                headers['User-Agent'] = self.get_random_user_agent()
                
                response = self.session.get(url, headers=headers, timeout=10)
                
                # Handle rate limiting
                if response.status_code == 429:  # Too Many Requests
                    wait_time = (2 ** attempt) + random.uniform(1, 3)
                    print(f"🚫 Rate limited (429), waiting {wait_time:.1f}s before retry {attempt + 1}/{max_retries}")
                    time.sleep(wait_time)
                    continue
                
                # Handle other HTTP errors
                if response.status_code >= 500:  # Server errors
                    wait_time = (2 ** attempt) + random.uniform(0.5, 1.5)
                    print(f"⚠️ Server error {response.status_code}, waiting {wait_time:.1f}s before retry")
                    time.sleep(wait_time)
                    continue
                
                return response
                
            except requests.exceptions.RequestException as e:
                if attempt == max_retries - 1:
                    raise e
                wait_time = (2 ** attempt) + random.uniform(0.5, 1.0)
                print(f"⚠️ Request error: {e}, waiting {wait_time:.1f}s before retry")
                time.sleep(wait_time)
        
        raise Exception(f"Failed to make request after {max_retries} attempts")
    
    def extract_emails_with_context(self, text):
        """Extract email addresses with their surrounding context"""
        if not text:
            return []
        
        # Split text into lines and sentences for context
        lines = text.split('\n')
        sentences = re.split(r'[.!?]+', text)
        
        found_emails = []
        
        # First, look for emails in individual lines
        for line in lines:
            line = line.strip()
            # FIX: Check for characters that indicate a potential email (standard OR obfuscated)
            line_lower = line.lower()
            if '@' in line_lower or 'at' in line_lower or '[' in line_lower:
                # Check if this line contains an email
                email_patterns = [
                    r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b',
                    r'\b[A-Za-z0-9._%+-]+\s*@\s*[A-Za-z0-9.-]+\s*\.\s*[A-Z|a-z]{2,}\b',
                    r'\b[A-Za-z0-9._%+-]+\[at\][A-Za-z0-9.-]+\[dot\][A-Z|a-z]{2,}\b',
                    r'\b[A-Za-z0-9._%+-]+\(at\)[A-Za-z0-9.-]+\(dot\)[A-Z|a-z]{2,}\b',
                    # New patterns to handle emojis and other characters before emails
                    r'[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}',
                    r'[A-Za-z0-9._%+-]+\s*@\s*[A-Za-z0-9.-]+\s*\.\s*[A-Z|a-z]{2,}',
                ]
                
                for pattern in email_patterns:
                    if re.search(pattern, line, re.IGNORECASE):
                        # Extract the actual email for validation
                        email_matches = re.findall(pattern, line, re.IGNORECASE)
                        for email_match in email_matches:
                            # Clean the email
                            clean_email = email_match.strip()
                            clean_email = re.sub(r'\s+', '', clean_email)
                            clean_email = clean_email.replace('[at]', '@').replace('(at)', '@')
                            clean_email = clean_email.replace('[.]', '.').replace('(dot)', '.')
                            clean_email = clean_email.replace('[dot]', '.').replace('(dot)', '.')
                            
                            # Validate the cleaned email
                            if re.match(r'^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}$', clean_email):
                                found_emails.append({
                                    'email': clean_email.lower(),
                                    'context': line.strip(),
                                    'type': 'line'
                                })
                        break
        
        # If no emails found in lines, try sentences
        if not found_emails:
            for sentence in sentences:
                sentence = sentence.strip()
                # FIX: Apply the same logic to the sentence check
                sentence_lower = sentence.lower()
                if ('@' in sentence_lower or 'at' in sentence_lower or '[' in sentence_lower) and len(sentence) < 500:
                    email_patterns = [
                        r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b',
                        r'\b[A-Za-z0-9._%+-]+\s*@\s*[A-Za-z0-9.-]+\s*\.\s*[A-Z|a-z]{2,}\b',
                        # New patterns to handle emojis and other characters before emails
                        r'[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}',
                        r'[A-Za-z0-9._%+-]+\s*@\s*[A-Za-z0-9.-]+\s*\.\s*[A-Z|a-z]{2,}',
                    ]
                    
                    for pattern in email_patterns:
                        if re.search(pattern, sentence, re.IGNORECASE):
                            email_matches = re.findall(pattern, sentence, re.IGNORECASE)
                            for email_match in email_matches:
                                clean_email = email_match.strip().lower()
                                if re.match(r'^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}$', clean_email):
                                    found_emails.append({
                                        'email': clean_email,
                                        'context': sentence.strip(),
                                        'type': 'sentence'
                                    })
                            break
        
        # Remove duplicates based on email address
        unique_emails = []
        seen_emails = set()
        for item in found_emails:
            if item['email'] not in seen_emails:
                unique_emails.append(item)
                seen_emails.add(item['email'])
        
        return unique_emails
    
    def extract_emails(self, text):
        """Legacy method - extract just email addresses for backward compatibility"""
        email_data = self.extract_emails_with_context(text)
        return [item['email'] for item in email_data]
    
    def scrape_with_requests(self, username):
        """Try to scrape using requests first (faster)"""
        try:
            # CPU MONITORING: Track CPU usage
            cpu_monitor.get_cpu_history()
            
            # Clean username
            username = username.replace('@', '').strip()
            url = f"https://www.tiktok.com/@{username}"
            print(f"🌐 REQUESTS: Fetching URL: {url}")
            
            # PERFORMANCE: Use shared session for connection reuse
            # Remove gzip from headers to avoid encoding issues
            headers = self.headers.copy()
            headers['Accept-Encoding'] = 'identity'  # No compression
            self.session.headers.update(headers)
            
            response = self.make_request_with_backoff(url, max_retries=3)
            print(f"🌐 REQUESTS: Response status code: {response.status_code}")
            if response.status_code == 200:
                # Get raw content and decode properly
                content = response.content
                
                # Try to detect encoding
                try:
                    import chardet
                    detected = chardet.detect(content)
                    encoding = detected.get('encoding', 'utf-8')
                except ImportError:
                    encoding = 'utf-8'
                
                try:
                    text_content = content.decode(encoding, errors='ignore')
                except:
                    text_content = content.decode('utf-8', errors='ignore')
                
                soup = BeautifulSoup(text_content, 'html.parser')
                
                # Updated selectors based on current TikTok structure
                bio_selectors = [
                    '[data-e2e="user-bio"]',
                    'h2[data-e2e="user-bio"]',
                    '[data-testid="user-bio"]',
                    '.css-1mf3iq5-H2ShareDesc',
                    '.tiktok-1mf3iq5-H2ShareDesc',
                    'h2.tiktok-1mf3iq5-H2ShareDesc',
                    '.user-bio',
                    '.profile-bio'
                ]
                
                bio_text = ""
                for selector in bio_selectors:
                    bio_elements = soup.select(selector)
                    for bio_element in bio_elements:
                        # FIXED: Use separator='\n' to preserve newlines like Selenium does
                        text = bio_element.get_text(separator='\n', strip=True)
                        if text and len(text) > 0:
                            bio_text = text
                            break
                    if bio_text:
                        break
                
                # Check for JSON data in various script tags
                if not bio_text:
                    scripts = soup.find_all('script')
                    for script in scripts:
                        if script.string and ('signature' in script.string or 'bio' in script.string.lower() or 'userInfo' in script.string):
                            try:
                                # Try to extract JSON data
                                script_content = script.string.strip()
                                
                                # Look for userInfo data that contains the signature/bio
                                if 'webapp.user-detail' in script_content:
                                    # Extract the userInfo data
                                    import re
                                    pattern = r'"signature":"([^"]*)"'
                                    match = re.search(pattern, script_content)
                                    if match:
                                        bio_text = match.group(1)
                                        # Decode escape sequences
                                        bio_text = bio_text.replace('\\n', '\n').replace('\\"', '"')
                                        print(f"🌐 REQUESTS: Found bio in JSON data: '{bio_text[:100]}{'...' if len(bio_text) > 100 else ''}'")
                                        break
                                
                                # Fallback to old method
                                if not bio_text and script_content.startswith('window.__INITIAL_STATE__'):
                                    json_str = script_content.replace('window.__INITIAL_STATE__=', '').rstrip(';')
                                    data = json.loads(json_str)
                                elif not bio_text and script_content.startswith('{') and script_content.endswith('}'):
                                    data = json.loads(script_content)
                                else:
                                    continue
                                
                                # Search for signature/bio in the JSON structure
                                if not bio_text:
                                    bio_text = self._extract_bio_from_json(data)
                                    if bio_text:
                                        break
                            except Exception as e:
                                print(f"🌐 REQUESTS: JSON parsing error: {e}")
                                continue
                
                # Get page text for login detection and fallback search
                page_text = soup.get_text()
                
                # Also search in all text content for email patterns as fallback
                if not bio_text:
                    # Look for potential bio sections in page text
                    lines = page_text.split('\n')
                    for i, line in enumerate(lines):
                        line = line.strip()
                        # Skip common non-bio text
                        if line in ['Signature (Required):', 'Sign up', 'Log in', 'Following', 'Followers', 'Likes']:
                            continue
                        if '@' in line and '.' in line and len(line) < 200:  # Potential bio with email
                            bio_text = line
                            break
                    
                    # If still no bio but we found "Signature (Required):", it means no bio set
                    if not bio_text and 'Signature (Required):' in page_text:
                        bio_text = "No bio set (Signature Required)"
                    
                # Check if we got the login page instead of profile
                if 'Make Your Day' in page_text and not bio_text:
                    bio_text = "TikTok_LOGIN_REQUIRED"
                    print(f"🌐 REQUESTS: Login page detected for {username}")
                
                print(f"🌐 REQUESTS: Bio extracted for {username}: '{bio_text[:100]}{'...' if len(bio_text) > 100 else ''}'")
                return bio_text
        except Exception as e:
            print(f"Requests method failed: {str(e)}")
            return None
    
    def _extract_bio_from_json(self, data):
        """Recursively search for bio/signature in JSON data"""
        if isinstance(data, dict):
            for key, value in data.items():
                if key.lower() in ['signature', 'bio', 'desc', 'description']:
                    if isinstance(value, str) and value.strip():
                        return value.strip()
                elif isinstance(value, (dict, list)):
                    result = self._extract_bio_from_json(value)
                    if result:
                        return result
        elif isinstance(data, list):
            for item in data:
                result = self._extract_bio_from_json(item)
                if result:
                    return result
        return None
    
# Playwright scraping method removed
    
    def scrape_with_selenium(self, username):
        """PERFORMANCE: Use WebDriver pool for Selenium scraping"""
        driver = None
        try:
            # CPU MONITORING: Track CPU usage
            cpu_monitor.get_cpu_history()
            
            # Check if WebDriver pool is available
            if not webdriver_pool:
                print("WebDriver pool not available, skipping Selenium method")
                return None
                
            username = username.replace('@', '').strip()
            url = f"https://www.tiktok.com/@{username}"
            
            # Get driver from pool
            driver = webdriver_pool.get_driver(timeout=5)
            if not driver:
                return None
            
            # Validate driver session before using
            try:
                driver.current_url  # Test if session is valid
            except Exception as e:
                print(f"Driver session invalid, creating new one: {e}")
                try:
                    driver.quit()
                except:
                    pass
                driver = webdriver_pool._create_driver()
                if not driver:
                    return None
            
            driver.get(url)
            print(f"🔧 SELENIUM: Loaded URL: {url}")

            # Reduced wait time for faster processing
            time.sleep(0.5)  # Optimized for maximum speed
            print(f"🔧 SELENIUM: Waiting for page to load...")
            
            # Try multiple selectors with different waiting strategies
            bio_selectors = [
                '[data-e2e="user-bio"]',
                'h2[data-e2e="user-bio"]',
                '[data-testid="user-bio"]',
                '.css-1mf3iq5-H2ShareDesc',
                '.tiktok-1mf3iq5-H2ShareDesc',
                'h2.tiktok-1mf3iq5-H2ShareDesc',
                '.user-bio',
                '.profile-bio',
                '[data-e2e="user-subtitle"]'
            ]
            
            bio_text = ""
            
            # First, try to find elements that are immediately available
            print(f"🔧 SELENIUM: Searching for bio elements with {len(bio_selectors)} selectors")
            for i, selector in enumerate(bio_selectors):
                try:
                    elements = driver.find_elements(By.CSS_SELECTOR, selector)
                    print(f"🔧 SELENIUM: Selector {i+1}/{len(bio_selectors)} '{selector}' found {len(elements)} elements")
                    for element in elements:
                        text = element.text.strip()
                        if text and len(text) > 0:
                            bio_text = text
                            print(f"🔧 SELENIUM: Found bio with selector '{selector}': '{text[:50]}{'...' if len(text) > 50 else ''}'")
                            break
                    if bio_text:
                        break
                except Exception as e:
                    print(f"🔧 SELENIUM: Error with selector '{selector}': {e}")
                    continue
            
            # If no bio found, try waiting for elements to appear (reduced timeout)
            if not bio_text:
                print(f"🔧 SELENIUM: No immediate bio found, trying WebDriverWait...")
                for i, selector in enumerate(bio_selectors):
                    try:
                        element = WebDriverWait(driver, 1).until(  # Optimized for speed
                            EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                        )
                        bio_text = element.text.strip()
                        if bio_text:
                            print(f"🔧 SELENIUM: Found bio with WebDriverWait selector '{selector}': '{bio_text[:50]}{'...' if len(bio_text) > 50 else ''}'")
                            break
                    except Exception as e:
                        print(f"🔧 SELENIUM: WebDriverWait failed for selector '{selector}': {e}")
                        continue
            
            # If still no bio, try to extract from page source
            if not bio_text:
                print(f"🔧 SELENIUM: No bio found with selectors, trying page source extraction...")
                page_source = driver.page_source
                soup = BeautifulSoup(page_source, 'html.parser')
                
                # Look for any text that might contain emails
                all_text = soup.get_text()
                lines = all_text.split('\n')
                print(f"🔧 SELENIUM: Searching through {len(lines)} lines of page source...")
                for line in lines:
                    line = line.strip()
                    if '@' in line and '.' in line and len(line) < 200:
                        # Check if it looks like a bio (contains email and reasonable length)
                        if re.search(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b', line):
                            bio_text = line
                            print(f"🔧 SELENIUM: Found bio in page source: '{bio_text[:50]}{'...' if len(bio_text) > 50 else ''}'")
                            break
            
            if bio_text:
                print(f"🔧 SELENIUM: Final bio result for {username}: '{bio_text[:100]}{'...' if len(bio_text) > 100 else ''}'")
            else:
                print(f"🔧 SELENIUM: No bio found for {username} with any method")
            
            return bio_text
            
        except Exception as e:
            print(f"Selenium method failed: {str(e)}")
            # If it's a session error, mark driver as invalid
            if "invalid session id" in str(e).lower() or "session" in str(e).lower():
                try:
                    if driver:
                        driver.quit()
                except:
                    pass
                driver = None  # Don't return invalid driver to pool
            return None
        finally:
            if driver:
                # Return driver to pool instead of closing
                webdriver_pool.return_driver(driver)
    
    def scrape_bio(self, username):
        """PERFORMANCE: Main method with caching and OPTIMIZED scraping order"""
        username = username.replace('@', '').strip().lower()
        
        print(f"🔍 Starting scrape for username: {username}")
        
        # 1. Check cache first (fastest)
        cached_result = profile_cache.get(username, expire_time=True)
        
        # cached_result is a (value, expire_time) tuple
        if cached_result is not None and cached_result[0] is not None and cached_result[1] is not None and cached_result[1] > time.time():
            print(f"✅ CACHE HIT for {username} - returning cached result")
            return cached_result[0].get('bio', '')  # [0] is the value
            
        print(f"❌ CACHE MISS for {username} - starting fresh scrape")
        
        bio = None
        method_used = None
        
        # 2. Try requests first (lightweight and very fast)
        print(f"🌐 Attempting requests method for {username}")
        try:
            bio = self.scrape_with_requests(username)
            if bio:
                method_used = "requests"
                print(f"✅ REQUESTS SUCCESS for {username} - bio length: {len(bio) if bio else 0}")
            else:
                print(f"⚠️ REQUESTS RETURNED EMPTY for {username}")
        except Exception as e:
            print(f"❌ REQUESTS FAILED for {username}: {e}")
            bio = None  # Ensure it's None so Selenium runs
            
        # 3. If requests failed or got a login page, escalate to Selenium
        #    (Selenium is the slow, heavy, but more reliable fallback)
        if not bio or bio == "TikTok_LOGIN_REQUIRED":
            if not bio:
                print(f"🚀 ESCALATING TO SELENIUM for {username} (requests returned empty)")
            else:
                print(f"🚀 ESCALATING TO SELENIUM for {username} (login page detected)")
                
            try:
                print(f"🔧 Attempting Selenium method for {username}")
                bio = self.scrape_with_selenium(username)
                if bio:
                    method_used = "selenium"
                    print(f"✅ SELENIUM SUCCESS for {username} - bio length: {len(bio) if bio else 0}")
                else:
                    print(f"⚠️ SELENIUM RETURNED EMPTY for {username}")
            except Exception as e:
                print(f"❌ SELENIUM FAILED for {username}: {e}")
                bio = None

        # 4. Final result logging
        if bio:
            print(f"🎉 FINAL SUCCESS for {username} using {method_used} method - bio: '{bio[:100]}{'...' if len(bio) > 100 else ''}'")
            
            # Check for emails in the bio
            email_data = self.extract_emails_with_context(bio)
            emails = [item['email'] for item in email_data]
            print(f"📧 EMAIL DETECTION for {username}: found {len(emails)} emails - {emails}")
        else:
            print(f"💥 FINAL FAILURE for {username} - both methods failed")

        # 5. Cache the result (even if it's None, to prevent re-scraping failures)
        cache_data = {
            'bio': bio,  # bio will be None if both methods failed
            'scraped_at': time.time(),
            'username': username
        }
        # set() is thread-safe and writes to disk/memory
        profile_cache.set(username, cache_data, expire=86400)  # 86400 seconds = 24 hours
        
        return bio
    
    def scrape_bio_force_refresh(self, username):
        """Scrape bio with cache bypass for fresh data"""
        username = username.replace('@', '').strip().lower()
        
        print(f"Force refresh for {username} - bypassing cache")
        
        bio = None
        
        # 2. Try requests first (lightweight and very fast)
        try:
            bio = self.scrape_with_requests(username)
        except Exception as e:
            print(f"Requests method failed outright: {e}")
            bio = None
            
        # 3. If requests failed or got a login page, escalate to Selenium
        if not bio or bio == "TikTok_LOGIN_REQUIRED":
            if not bio:
                print(f"Requests failed for {username}. Escalating to Selenium.")
            else:
                print(f"Requests got login page for {username}. Escalating to Selenium.")
                
            try:
                bio = self.scrape_with_selenium(username)
            except Exception as e:
                print(f"Selenium method also failed: {e}")
                bio = None

        # Cache the result for future use (even with force refresh, we still cache the new result)
        cache_data = {
            'bio': bio,
            'scraped_at': time.time(),
            'username': username
        }
        profile_cache.set(username, cache_data, expire=86400)
        
        return bio

# PERFORMANCE: Concurrent bulk processing function with memory optimization and CPU throttling
def process_username_batch(usernames, max_workers=20, force_refresh=False):
    """Process multiple usernames concurrently with memory optimization and CPU throttling"""
    results = []
    # PERFORMANCE: Create shared scraper instance for session reuse
    scraper = TikTokScraper()
    
    # RATE LIMITING: Process in smaller batches to avoid detection
    batch_size = 25  # Process 25 usernames per batch
    
    # CPU MONITORING: Calculate optimal workers based on CPU cores
    optimal_workers = cpu_monitor.get_optimal_workers()
    print(f"🔄 Using {optimal_workers} workers (limited to 20 for resource management)")
    
    def process_single_username(username):
        print(f"📋 PROCESSING USERNAME: {username}")
        try:
            # CPU MONITORING: Track CPU usage
            cpu_monitor.get_cpu_history()
            
            # Use force_refresh to bypass cache if requested
            if force_refresh:
                print(f"🔄 FORCE REFRESH enabled for {username}")
                bio = scraper.scrape_bio_force_refresh(username)
            else:
                bio = scraper.scrape_bio(username)
            if bio:
                if bio == "TikTok_LOGIN_REQUIRED":
                    return {
                        'username': username,
                        'success': False,
                        'error': 'TikTok requires login to view profiles',
                        'suggestion': f'Manual URL: https://www.tiktok.com/@{username}'
                    }
                
                # MEMORY OPTIMIZATION: Extract emails without storing full bio in memory
                email_data = scraper.extract_emails_with_context(bio)
                emails = [item['email'] for item in email_data]
                
                # MEMORY OPTIMIZATION: Only store essential data
                result = {
                    'username': username,
                    'success': True,
                    'bio': bio,  # Keep bio for now, but could be optimized further
                    'emails': emails,
                    'email_data': email_data,
                    'email_count': len(emails)
                }
                
                print(f"✅ USERNAME {username} COMPLETED SUCCESSFULLY - Found {len(emails)} emails: {emails}")
                
                # MEMORY OPTIMIZATION: Clear bio from memory after processing
                bio = None
                return result
            else:
                print(f"❌ USERNAME {username} FAILED - Could not retrieve bio")
                return {
                    'username': username,
                    'success': False,
                    'error': 'Could not retrieve bio'
                }
        except Exception as e:
            print(f"💥 USERNAME {username} EXCEPTION: {str(e)}")
            return {
                'username': username,
                'success': False,
                'error': f'Error: {str(e)}'
            }
    
    # RATE LIMITING: Process usernames in batches with breaks
    total_batches = (len(usernames) + batch_size - 1) // batch_size
    print(f"📊 Processing {len(usernames)} usernames in {total_batches} batches of {batch_size}")
    
    for batch_num in range(total_batches):
        start_idx = batch_num * batch_size
        end_idx = min(start_idx + batch_size, len(usernames))
        batch_usernames = usernames[start_idx:end_idx]
        
        print(f"🔄 Processing batch {batch_num + 1}/{total_batches} ({len(batch_usernames)} usernames)")
        
        # Process current batch
        with ThreadPoolExecutor(max_workers=optimal_workers) as executor:
            # Submit batch tasks
            future_to_username = {executor.submit(process_single_username, username): username for username in batch_usernames}
            
            # Collect results as they complete
            batch_completed = 0
            for future in as_completed(future_to_username):
                username = future_to_username[future]
                batch_completed += 1
                total_completed = start_idx + batch_completed
                print(f"📊 COMPLETED {total_completed}/{len(usernames)}: {username}")
                try:
                    result = future.result()
                    results.append(result)
                    print(f"✅ RESULT ADDED for {username}: success={result.get('success', False)}")
                        
                except Exception as e:
                    error_result = {
                        'username': username,
                        'success': False,
                        'error': f'Processing error: {str(e)}'
                    }
                    results.append(error_result)
                    print(f"❌ ERROR RESULT ADDED for {username}: {str(e)}")
        
        # Continue to next batch immediately (no rate limiting break)
    
    print(f"📊 BULK PROCESSING COMPLETE: {len(results)} results collected from {len(usernames)} usernames")
    return results

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/scrape', methods=['POST'])
def scrape_bio():
    try:
        data = request.get_json()
        username = data.get('username', '').strip()
        
        if not username:
            return jsonify({'error': 'Please provide a TikTok username'})
        
        print(f"Scraping bio for username: {username}")  # Debug log
        
        scraper = TikTokScraper()
        bio = scraper.scrape_bio(username)
        
        print(f"Retrieved bio: {bio[:100] if bio else 'None'}...")  # Debug log
        
        if bio:
            # Check if TikTok is requiring login
            if bio == "TikTok_LOGIN_REQUIRED":
                return jsonify({
                    'success': False,
                    'error': 'TikTok requires login to view profiles. This is a limitation of automated scraping. Try: 1) Visit the profile manually in your browser, 2) Use TikTok\'s official API, or 3) Try a different approach.',
                    'suggestion': f'Manual URL: https://www.tiktok.com/@{username}'
                })
            
            # Get emails with context
            email_data = scraper.extract_emails_with_context(bio)
            emails = [item['email'] for item in email_data]
            print(f"Extracted emails with context: {email_data}")  # Debug log
            
            return jsonify({
                'success': True,
                'username': username,
                'bio': bio,
                'emails': emails,
                'email_data': email_data,  # Include context data
                'email_count': len(emails)
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Could not retrieve bio. User might not exist, profile might be private, or TikTok is blocking requests. Try again in a few moments.'
            })
    
    except Exception as e:
        print(f"Error in scrape_bio: {str(e)}")  # Debug log
        return jsonify({
            'success': False,
            'error': f'An error occurred: {str(e)}'
        })

@app.route('/bulk-scrape', methods=['POST'])
def bulk_scrape():
    """PERFORMANCE: Concurrent bulk processing endpoint"""
    try:
        data = request.get_json()
        usernames = data.get('usernames', [])
        force_refresh = data.get('force_refresh', False)  # Cache-busting parameter
        
        if not usernames:
            return jsonify({'error': 'Please provide usernames'})
        
        # Process usernames in batches for better performance and memory management
        BATCH_SIZE = 1000
        if len(usernames) > BATCH_SIZE:
            return jsonify({
                'error': f'Too many usernames for single batch. Processing {len(usernames)} usernames in batches of {BATCH_SIZE}.',
                'batch_size': BATCH_SIZE,
                'total_usernames': len(usernames),
                'suggestion': f'Please split your request into batches of {BATCH_SIZE} usernames or fewer.'
            })
        
        print(f"Processing {len(usernames)} usernames concurrently...")
        if force_refresh:
            print("🔄 Force refresh enabled - bypassing cache for fresh data")
        
        # Process concurrently (3-5x faster than sequential)
        results = process_username_batch(usernames, max_workers=20, force_refresh=force_refresh)
        
        # Format results for frontend
        processed_results = []
        for result in results:
            if result['success']:
                processed_results.append({
                    'username': result['username'],
                    'success': True,
                    'emails': result['emails'],
                    'email_data': result['email_data'],
                    'bio': result['bio']
                })
            else:
                processed_results.append({
                    'username': result['username'],
                    'success': False,
                    'error': result['error'],
                    'suggestion': result.get('suggestion')
                })
        
        return jsonify({
            'success': True,
            'results': processed_results,
            'total': len(results),
            'successful': len([r for r in results if r['success']])
        })
        
    except Exception as e:
        print(f"Error in bulk_scrape: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Bulk processing failed: {str(e)}'
        })

@app.route('/export-excel', methods=['POST'])
def export_excel():
    try:
        data = request.get_json()
        results = data.get('results', [])
        
        # MEMORY OPTIMIZATION: Use generator instead of building full list
        def email_results_generator():
            for result in results:
                if result.get('emails') and len(result['emails']) > 0:
                    for email in result['emails']:
                        yield {
                            'username': result['username'],
                            'profile_url': result.get('profile_url', f"https://www.tiktok.com/@{result['username']}"),
                            'email': email
                        }
        
        # Count total results first
        email_count = sum(1 for _ in email_results_generator())
        if email_count == 0:
            return jsonify({'error': 'No email results to export'})
        
        # MEMORY OPTIMIZATION: Create workbook and write data directly (streaming)
        wb = Workbook()
        ws = wb.active
        ws.title = "TikTok Email Results"
        
        # Set headers
        headers = ['Name', 'Email']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
        
        # MEMORY OPTIMIZATION: Write data row by row without storing in memory
        row_num = 2
        for result in email_results_generator():
            # Name column (username)
            ws.cell(row=row_num, column=1, value=result['username'])
            
            # Email column
            ws.cell(row=row_num, column=2, value=result['email'])
            
            row_num += 1
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"tiktok_emails_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'error': f'Export failed: {str(e)}'})

@app.route('/clear-cache', methods=['POST'])
def clear_cache():
    try:
        count = profile_cache.clear()  # This clears both memory and disk
        return jsonify({
            'success': True,
            'message': f'Cache cleared successfully! {count} items removed.',
            'items_removed': count
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Failed to clear cache: {str(e)}'
        })

@app.route('/cache-stats', methods=['GET'])
def cache_stats():
    try:
        # Note: diskcache doesn't easily separate memory/file counts
        # It's one unified cache.
        item_count = len(profile_cache)
        cache_size_bytes = profile_cache.volume()
        cache_size_mb = round(cache_size_bytes / (1024 * 1024), 2)
        
        return jsonify({
            'success': True,
            'total_items': item_count,
            'total_cache_size_mb': cache_size_mb,
            'cache_directory': profile_cache.directory
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Failed to get cache stats: {str(e)}'
        })

@app.route('/system-stats', methods=['GET'])
def system_stats():
    """Get system performance statistics including CPU usage"""
    try:
        # Get CPU usage
        cpu_percent = cpu_monitor.get_cpu_usage()
        cpu_history = cpu_monitor.get_cpu_history()
        is_throttling = False  # No throttling anymore
        
        # Get memory usage
        memory = psutil.virtual_memory()
        memory_percent = memory.percent
        memory_available_gb = round(memory.available / (1024**3), 2)
        memory_used_gb = round(memory.used / (1024**3), 2)
        memory_total_gb = round(memory.total / (1024**3), 2)
        
        # Get disk usage
        disk = psutil.disk_usage('/')
        disk_percent = disk.percent
        disk_free_gb = round(disk.free / (1024**3), 2)
        disk_used_gb = round(disk.used / (1024**3), 2)
        disk_total_gb = round(disk.total / (1024**3), 2)
        
        # Get process count
        process_count = len(psutil.pids())
        
        return jsonify({
            'success': True,
            'cpu': {
                'current_percent': cpu_percent,
                'history': cpu_history,
                'is_throttling': is_throttling,
                'max_threshold': 100  # No throttling limit
            },
            'memory': {
                'percent': memory_percent,
                'available_gb': memory_available_gb,
                'used_gb': memory_used_gb,
                'total_gb': memory_total_gb
            },
            'disk': {
                'percent': disk_percent,
                'free_gb': disk_free_gb,
                'used_gb': disk_used_gb,
                'total_gb': disk_total_gb
            },
            'system': {
                'process_count': process_count,
                'uptime_seconds': time.time() - psutil.boot_time()
            }
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Failed to get system stats: {str(e)}'
        })

# CPU configuration endpoint removed - no throttling anymore

@app.route('/remove-success', methods=['POST'])
def remove_success():
    """Remove all successfully scraped usernames, keeping only failed/no email results"""
    try:
        data = request.get_json()
        results = data.get('results', [])
        
        # Filter out successful results (those with emails)
        filtered_results = []
        for result in results:
            # Keep results that have no emails or failed status
            if not result.get('emails') or len(result['emails']) == 0 or result.get('status') == 'failed':
                filtered_results.append(result)
        
        return jsonify({
            'success': True,
            'results': filtered_results,
            'removed_count': len(results) - len(filtered_results),
            'remaining_count': len(filtered_results),
            'message': f'Removed {len(results) - len(filtered_results)} successful results, kept {len(filtered_results)} failed/no email results'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Failed to remove successful results: {str(e)}'
        })

@app.route('/import-excel', methods=['POST'])
def import_excel():
    try:
        # Check if file was uploaded
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'})
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'})
        
        # Check file extension
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Invalid file format. Please upload an Excel file (.xlsx or .xls)'})
        
        # Save uploaded file temporarily
        filename = secure_filename(file.filename)
        temp_path = os.path.join(tempfile.gettempdir(), filename)
        file.save(temp_path)
        
        try:
            # Load the workbook
            wb = load_workbook(temp_path)
            ws = wb.active
            
            # Find the data rows (skip title and header rows)
            results = []
            header_found = False
            
            for row in ws.iter_rows(min_row=1, values_only=True):
                if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                    continue
                
                # Skip title row
                if not header_found and row[0] and 'TikTok' in str(row[0]):
                    continue
                
                # Look for header row
                if not header_found and row[0] and ('Username' in str(row[0]) or '@' in str(row[0])):
                    header_found = True
                    continue
                
                # Process data rows
                if header_found and len(row) >= 3:
                    username = str(row[0]).strip() if row[0] else ''
                    profile_url = str(row[1]).strip() if row[1] else ''
                    email = str(row[2]).strip() if row[2] else ''
                    
                    # Clean username (remove @)
                    if username.startswith('@'):
                        username = username[1:]
                    
                    # Skip empty rows
                    if not username:
                        continue
                    
                    # Create result object
                    result = {
                        'username': username,
                        'profile_url': profile_url if profile_url else f'https://www.tiktok.com/@{username}',
                        'emails': [email] if email and email != '❌ No email found' else [],
                        'email_data': [{'email': email, 'context': 'Imported from Excel', 'type': 'imported'}] if email and email != '❌ No email found' else [],
                        'status': 'success' if email and email != '❌ No email found' else 'failed'
                    }
                    
                    results.append(result)
            
            # Clean up temporary file
            os.remove(temp_path)
            
            if not results:
                return jsonify({'error': 'No valid data found in Excel file. Make sure it has Username, Profile URL, and Email columns.'})
            
            return jsonify({
                'success': True,
                'results': results,
                'count': len(results),
                'message': f'Successfully imported {len(results)} records'
            })
            
        except Exception as e:
            # Clean up temporary file on error
            if os.path.exists(temp_path):
                os.remove(temp_path)
            return jsonify({'error': f'Error reading Excel file: {str(e)}'})
        
    except Exception as e:
        return jsonify({'error': f'Import failed: {str(e)}'})

@app.teardown_appcontext
def cleanup_resources(error):
    """Enhanced cleanup on app shutdown"""
    try:
        if webdriver_pool:
            webdriver_pool.close_all()
    except Exception as e:
        print(f"Cleanup error: {e}")

import atexit
import signal
import os
import sys

def cleanup_on_exit():
    """Enhanced cleanup on exit"""
    try:
        print("\n🔄 Cleaning up processes...")
        
        # Close WebDriver pool
        if webdriver_pool:
            webdriver_pool.close_all()
            print("✅ WebDriver pool cleaned up")
        
        # Kill all Chrome processes related to this app
        try:
            import psutil
            current_pid = os.getpid()
            for proc in psutil.process_iter(['pid', 'name', 'cmdline', 'ppid']):
                try:
                    if proc.info['name'] and 'chrome' in proc.info['name'].lower():
                        cmdline = proc.info.get('cmdline', [])
                        # Check if it's a headless Chrome from our app
                        if cmdline and any('headless' in str(arg).lower() for arg in cmdline):
                            proc.terminate()
                            print(f"✅ Killed Chrome process: {proc.info['pid']}")
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    pass
        except ImportError:
            pass
        
        # Force garbage collection
        import gc
        gc.collect()
        print("✅ Cleanup completed")
        
    except Exception as e:
        print(f"❌ Exit cleanup error: {e}")

def signal_handler(signum, frame):
    """Handle termination signals"""
    print(f"\n🛑 Received signal {signum}, shutting down gracefully...")
    cleanup_on_exit()
    sys.exit(0)

# Register signal handlers
signal.signal(signal.SIGINT, signal_handler)   # Ctrl+C
signal.signal(signal.SIGTERM, signal_handler)  # Termination signal
atexit.register(cleanup_on_exit)

if __name__ == '__main__':
    import sys
    
    # Create a new process group to ensure all child processes are killed together
    try:
        if os.name == 'nt':  # Windows
            # On Windows, create a new process group
            import ctypes
            kernel32 = ctypes.windll.kernel32
            kernel32.SetConsoleCtrlHandler(None, False)
            kernel32.SetConsoleCtrlHandler(lambda signal: cleanup_on_exit() or True, True)
        else:  # Unix-like systems
            os.setpgrp()  # Create new process group
    except Exception as e:
        print(f"Warning: Could not set up process group: {e}")
    
    # Get port from command line argument or environment variable
    if len(sys.argv) > 1:
        try:
            port = int(sys.argv[1])
        except ValueError:
            print("Invalid port number. Using default port 5001")
            port = 5001
    else:
        port = int(os.environ.get('PORT', 5001))
    
    # Get host from command line or environment variable
    if len(sys.argv) > 2:
        host = sys.argv[2]
    else:
        host = os.environ.get('HOST', '0.0.0.0')  # Default to 0.0.0.0 to bind to all interfaces
    
    print(f"🚀 Starting TikTok Scraper on {host}:{port}")
    print("📋 Available ports: 5001, 5002, 5003")
    print("🌐 Available hosts: 127.0.0.1, 172.x.x.x, 0.0.0.0 (all interfaces)")
    print("💡 Usage: python app.py [port] [host] or python app.py 5002 172.16.0.1")
    print("⚠️  Press Ctrl+C to stop the server (will kill all child processes)")
    print("-" * 60)
    
    try:
        # PERFORMANCE: Configure Flask with optimized threading
        app.run(
            debug=False,  # Disable debug to prevent reloader process issues
            host=host, 
            port=port,
            use_reloader=False,  # Disable reloader to prevent orphaned processes
            threaded=True,  # Enable threading for concurrent requests
            processes=1  # Use single process to avoid threading issues
        )
    except KeyboardInterrupt:
        print("\n🛑 Keyboard interrupt received, shutting down gracefully...")
        cleanup_on_exit()
    except Exception as e:
        print(f"❌ Error starting server: {e}")
        cleanup_on_exit()
    finally:
        cleanup_on_exit()

